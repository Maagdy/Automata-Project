import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_analysis(results, output_file="results_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    headers = ["Word", "Type", "Line", "Column", "Word #"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Data rows
    for r in results:
        ws.append([
            r["word"],
            r["type"],
            r["line"],
            r["col"],
            r["word_number"]
        ])

    wb.save(output_file)
    print(f"Analysis saved to {output_file}")


def mask_phones(text, phones):
    masked_text = text

    for phone in phones:
        digits = re.sub(r"\D", "", phone)

        if len(digits) >= 6:
            masked = phone
            for i in range(6):
                digit = digits[-(i + 1)]
                masked = masked[::-1].replace(digit, "*", 1)[::-1]

            masked_text = masked_text.replace(phone, masked)

    return masked_text


def write_masked(text, phones, output_file="masked_output.xlsx"):
    masked = mask_phones(text, phones)

    wb = Workbook()
    ws = wb.active
    ws.title = "Masked Text"

    for line in masked.splitlines():
        ws.append([line])

    ws.column_dimensions["A"].width = 120

    wb.save(output_file)
    print(f"Masked text saved to {output_file}")
